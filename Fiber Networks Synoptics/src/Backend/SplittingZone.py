# Splitting Zone or JSO Zone class file
# from Cable import Cable
import re

from JSO import JSO
import ezdxf
import random
import openpyxl
import AttributeProcessing as ap


class SplittingZone:

    def __init__(self, entity, path: str):  # Instance Variables definition
        self.jso = JSO(entity)
        self.index = entity.get_attrib_text("TIPO")
        self.jfo_list = list()
        self.pdo_list = list()
        self.document = ezdxf.new('R2018')
        self.model_space = self.document.modelspace()
        self.link_table = self.link_table(path)
        self.coupled = dict()

    def add_pdo(self, pdo):
        if pdo not in self.pdo_list:
            self.pdo_list.append(pdo)

    def add_jfo(self, jfo):
        if jfo not in self.jfo_list:
            self.jfo_list.append(jfo)

    def is_pdo(self, ident):
        for pdo in self.pdo_list:
            if pdo.identifier == ident:
                return True
        return False

    def is_jfo(self, ident):
        for jfo in self.pdo_list:
            if jfo.identifier == ident:
                return True
        return False

    def add_coupled(self, entity, index):
        if entity.get_attrib_text("TIPO_ESTRUTURA") == "ACOPLADA":
            self.coupled["PDO" + entity.get_attrib_text("JFO_PDO")] = index

    @staticmethod
    def get_random_point():
        """Returns random x, y coordinates."""
        x = random.randint(-100, 100)
        y = random.randint(-100, 100)
        return x, y

    def to_string(self) -> str:
        result = "\tIndex: " + self.index + "\n" + "\tJSO: " + self.jso.entity.get_attrib_text("TIPO") + "\n"
        for pdo in self.pdo_list:
            result += "\t\tPDO" + pdo.identifier + "\n"
        for jfo in self.jfo_list:
            result += "\t\t" + jfo.identifier + "\n"
        return result

    def create_synoptics(self):
        placing_points = [self.get_random_point() for _ in range(len(self.pdo_list))]
        block = self.document.blocks.new(name='PDO_SIN')
        i = 0
        for pdo in self.pdo_list:
            entity = pdo.entity
            attribs = entity.attribs
            self.model_space.add_auto_blockref('PDO_SIN', placing_points[i], values=attribs, dxfattribs={
                'xscale': entity.dxf.xscale,
                'yscale': entity.dxf.yscale,
                'rotation': 0
            })
            i += 1
        for flag_ref in self.model_space.query('INSERT[name=="PDO_SIN"]'):
            print("FLAG REFERENCE: " + str(flag_ref))
        filename = r"C:\Users\thech\Desktop\MIEI\Dissertation\Fiber Networks Synoptics\\" + self.index + ".dxf"
        # self.document.saveas(filename)

    @staticmethod
    def sort_branch(zone_dict: dict, key) -> dict:
        result = {}
        for k in zone_dict[key].keys():
            if k == key:
                continue
        return result

    def sort_map(self, zone_dict: dict) -> dict:
        result = {}
        for key in zone_dict[self.index].keys():
            if key == self.index:
                continue
            result[key] = self.sort_branch(zone_dict, key)
        return result

    def link_table(self, path) -> dict:
        row_start = 13
        # filename = r"C:\Users\thech\Desktop\MIEI\Dissertation\Fiber Networks Synoptics\doc\dxf\exemplo VODAFONE\04 - Tabelas de juntas de Rede Secundária\Tabelas de Ligação da " + self.index + ".xlsx"
        # filename = r"C:\Users\User\Desktop\Dissertation\Fiber Networks Synoptics\doc\dxf\exemplo VODAFONE\04 - Tabelas de juntas de Rede Secundária\Tabelas de Ligação da " + self.index + ".xlsx"
        filename = path + "\\Tabelas de Ligação da " + self.index + ".xlsx"
        workbook = openpyxl.load_workbook(filename)
        zone_dict = {}
        for sheet in workbook.worksheets:
            if sheet.title == "ESQ":
                continue
            equipment_dict = {}
            for row in range(row_start, sheet.max_row):
                line = []
                for col in sheet.iter_cols(1, 14):
                    line.append(col[row].value)
                if all(item is None for item in line):
                    continue
                if line[0] == '' or line[13] == '' or line[7] == '':
                    continue
                equipment_dict[line[13]] = []
                if line[7] and line[13] != sheet.title and re.match('^VDF', line[7]):
                    equipment_dict[line[13]].append(line[7])
                elif line[0] and line[13] == sheet.title and re.match('^VDF', line[0]):
                    equipment_dict[line[13]].append(line[0])
                if len(equipment_dict[line[13]]) > 0:
                    zone_dict[sheet.title] = equipment_dict
        return zone_dict
